from openpyxl import workbook, load_workbook

application_path = os.path.dirname(sys.executable)

print('Warning')

print('This software is just made to update the Return analysis sheet quikly. Please do not add or remove either coloumns or rows in the sheet manually')

print('If you need to add or remove columns and rows, Please contact Mohamed')

print('To get started, please type s in small case to begin')


driver = input('Enter s to start ')

while driver == 's':
    route = int(input('What is the route number? '))

    if route == 101:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E2'].value = total
        ws['F2'].value = dex08
        ws['G2'].value = dex07
        ws['H2'].value = dex03
        ws['I2'].value = dex05
        ws['J2'].value = dex17
        ws['K2'].value = dex01
        ws['M2'].value = novan
        ws['N2'].value = colstop
        ws['O2'].value = colpux
        ws['P2'].value = retn
        ws['Q2'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 102:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E2'].value = total
        ws['F3'].value = dex08
        ws['G3'].value = dex07
        ws['H3'].value = dex03
        ws['I3'].value = dex05
        ws['J3'].value = dex17
        ws['K3'].value = dex01
        ws['M3'].value = novan
        ws['N3'].value = colstop
        ws['O3'].value = colpux
        ws['P3'].value = retn
        ws['Q3'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 103:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E4'].value = total
        ws['F4'].value = dex08
        ws['G4'].value = dex07
        ws['H4'].value = dex03
        ws['I4'].value = dex05
        ws['J4'].value = dex17
        ws['K4'].value = dex01
        ws['M4'].value = novan
        ws['N4'].value = colstop
        ws['O4'].value = colpux
        ws['P4'].value = retn
        ws['Q4'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 104:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E5'].value = total
        ws['F5'].value = dex08
        ws['G5'].value = dex07
        ws['H5'].value = dex03
        ws['I5'].value = dex05
        ws['J5'].value = dex17
        ws['K5'].value = dex01
        ws['M5'].value = novan
        ws['N5'].value = colstop
        ws['O5'].value = colpux
        ws['P5'].value = retn
        ws['Q5'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 105:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E6'].value = total
        ws['F6'].value = dex08
        ws['G6'].value = dex07
        ws['H6'].value = dex03
        ws['I6'].value = dex05
        ws['J6'].value = dex17
        ws['K6'].value = dex01
        ws['M6'].value = novan
        ws['N6'].value = colstop
        ws['O6'].value = colpux
        ws['P6'].value = retn
        ws['Q6'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 106:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E7'].value = total
        ws['F7'].value = dex08
        ws['G7'].value = dex07
        ws['H7'].value = dex03
        ws['I7'].value = dex05
        ws['J7'].value = dex17
        ws['K7'].value = dex01
        ws['M7'].value = novan
        ws['N7'].value = colstop
        ws['O7'].value = colpux
        ws['P7'].value = retn
        ws['Q7'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 107:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E8'].value = total
        ws['F8'].value = dex08
        ws['G8'].value = dex07
        ws['H8'].value = dex03
        ws['I8'].value = dex05
        ws['J8'].value = dex17
        ws['K8'].value = dex01
        ws['M8'].value = novan
        ws['N8'].value = colstop
        ws['O8'].value = colpux
        ws['P8'].value = retn
        ws['Q8'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 108:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E9'].value = total
        ws['F9'].value = dex08
        ws['G9'].value = dex07
        ws['H9'].value = dex03
        ws['I9'].value = dex05
        ws['J9'].value = dex17
        ws['K9'].value = dex01
        ws['M9'].value = novan
        ws['N9'].value = colstop
        ws['O9'].value = colpux
        ws['P9'].value = retn
        ws['Q9'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 109:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E10'].value = total
        ws['F10'].value = dex08
        ws['G10'].value = dex07
        ws['H10'].value = dex03
        ws['I10'].value = dex05
        ws['J10'].value = dex17
        ws['K10'].value = dex01
        ws['M10'].value = novan
        ws['N10'].value = colstop
        ws['O10'].value = colpux
        ws['P10'].value = retn
        ws['Q10'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 110:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E11'].value = total
        ws['F11'].value = dex08
        ws['G11'].value = dex07
        ws['H11'].value = dex03
        ws['I11'].value = dex05
        ws['J11'].value = dex17
        ws['K11'].value = dex01
        ws['M11'].value = novan
        ws['N11'].value = colstop
        ws['O11'].value = colpux
        ws['P11'].value = retn
        ws['Q11'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 111:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E12'].value = total
        ws['F12'].value = dex08
        ws['G12'].value = dex07
        ws['H12'].value = dex03
        ws['I12'].value = dex05
        ws['J12'].value = dex17
        ws['K12'].value = dex01
        ws['M12'].value = novan
        ws['N12'].value = colstop
        ws['O12'].value = colpux
        ws['P12'].value = retn
        ws['Q12'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 121:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E19'].value = total
        ws['F19'].value = dex08
        ws['G19'].value = dex07
        ws['H19'].value = dex03
        ws['I19'].value = dex05
        ws['J19'].value = dex17
        ws['K19'].value = dex01
        ws['M19'].value = novan
        ws['N19'].value = colstop
        ws['O19'].value = colpux
        ws['P19'].value = retn
        ws['Q19'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 122:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E20'].value = total
        ws['F20'].value = dex08
        ws['G20'].value = dex07
        ws['H20'].value = dex03
        ws['I20'].value = dex05
        ws['J20'].value = dex17
        ws['K20'].value = dex01
        ws['M20'].value = novan
        ws['N20'].value = colstop
        ws['O20'].value = colpux
        ws['P20'].value = retn
        ws['Q20'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 123:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E21'].value = total
        ws['F21'].value = dex08
        ws['G21'].value = dex07
        ws['H21'].value = dex03
        ws['I21'].value = dex05
        ws['J21'].value = dex17
        ws['K21'].value = dex01
        ws['M21'].value = novan
        ws['N21'].value = colstop
        ws['O21'].value = colpux
        ws['P21'].value = retn
        ws['Q21'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 124:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E22'].value = total
        ws['F22'].value = dex08
        ws['G22'].value = dex07
        ws['H22'].value = dex03
        ws['I22'].value = dex05
        ws['J22'].value = dex17
        ws['K22'].value = dex01
        ws['M22'].value = novan
        ws['N22'].value = colstop
        ws['O22'].value = colpux
        ws['P22'].value = retn
        ws['Q22'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 125:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E23'].value = total
        ws['F23'].value = dex08
        ws['G23'].value = dex07
        ws['H23'].value = dex03
        ws['I23'].value = dex05
        ws['J23'].value = dex17
        ws['K23'].value = dex01
        ws['M23'].value = novan
        ws['N23'].value = colstop
        ws['O23'].value = colpux
        ws['P23'].value = retn
        ws['Q23'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 126:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E24'].value = total
        ws['F24'].value = dex08
        ws['G24'].value = dex07
        ws['H24'].value = dex03
        ws['I24'].value = dex05
        ws['J24'].value = dex17
        ws['K24'].value = dex01
        ws['M24'].value = novan
        ws['N24'].value = colstop
        ws['O24'].value = colpux
        ws['P24'].value = retn
        ws['Q24'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 127:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E25'].value = total
        ws['F25'].value = dex08
        ws['G25'].value = dex07
        ws['H25'].value = dex03
        ws['I25'].value = dex05
        ws['J25'].value = dex17
        ws['K25'].value = dex01
        ws['M25'].value = novan
        ws['N25'].value = colstop
        ws['O25'].value = colpux
        ws['P25'].value = retn
        ws['Q25'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 128:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E26'].value = total
        ws['F26'].value = dex08
        ws['G26'].value = dex07
        ws['H26'].value = dex03
        ws['I26'].value = dex05
        ws['J26'].value = dex17
        ws['K26'].value = dex01
        ws['M26'].value = novan
        ws['N26'].value = colstop
        ws['O26'].value = colpux
        ws['P26'].value = retn
        ws['Q26'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 442:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E27'].value = total
        ws['F27'].value = dex08
        ws['G27'].value = dex07
        ws['H27'].value = dex03
        ws['I27'].value = dex05
        ws['J27'].value = dex17
        ws['K27'].value = dex01
        ws['M27'].value = novan
        ws['N27'].value = colstop
        ws['O27'].value = colpux
        ws['P27'].value = retn
        ws['Q27'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 141:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E34'].value = total
        ws['F34'].value = dex08
        ws['G34'].value = dex07
        ws['H34'].value = dex03
        ws['I34'].value = dex05
        ws['J34'].value = dex17
        ws['K34'].value = dex01
        ws['M34'].value = novan
        ws['N34'].value = colstop
        ws['O34'].value = colpux
        ws['P34'].value = retn
        ws['Q34'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 142:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E35'].value = total
        ws['F35'].value = dex08
        ws['G35'].value = dex07
        ws['H35'].value = dex03
        ws['I35'].value = dex05
        ws['J35'].value = dex17
        ws['K35'].value = dex01
        ws['M35'].value = novan
        ws['N35'].value = colstop
        ws['O35'].value = colpux
        ws['P35'].value = retn
        ws['Q35'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 143:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E36'].value = total
        ws['F36'].value = dex08
        ws['G36'].value = dex07
        ws['H36'].value = dex03
        ws['I36'].value = dex05
        ws['J36'].value = dex17
        ws['K36'].value = dex01
        ws['M36'].value = novan
        ws['N36'].value = colstop
        ws['O36'].value = colpux
        ws['P36'].value = retn
        ws['Q36'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 144:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E37'].value = total
        ws['F37'].value = dex08
        ws['G37'].value = dex07
        ws['H37'].value = dex03
        ws['I37'].value = dex05
        ws['J37'].value = dex17
        ws['K37'].value = dex01
        ws['M37'].value = novan
        ws['N37'].value = colstop
        ws['O37'].value = colpux
        ws['P37'].value = retn
        ws['Q37'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 145:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E38'].value = total
        ws['F38'].value = dex08
        ws['G38'].value = dex07
        ws['H38'].value = dex03
        ws['I38'].value = dex05
        ws['J38'].value = dex17
        ws['K38'].value = dex01
        ws['M38'].value = novan
        ws['N38'].value = colstop
        ws['O38'].value = colpux
        ws['P38'].value = retn
        ws['Q38'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 146:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E39'].value = total
        ws['F39'].value = dex08
        ws['G39'].value = dex07
        ws['H39'].value = dex03
        ws['I39'].value = dex05
        ws['J39'].value = dex17
        ws['K39'].value = dex01
        ws['M39'].value = novan
        ws['N39'].value = colstop
        ws['O39'].value = colpux
        ws['P39'].value = retn
        ws['Q39'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 448:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E40'].value = total
        ws['F40'].value = dex08
        ws['G40'].value = dex07
        ws['H40'].value = dex03
        ws['I40'].value = dex05
        ws['J40'].value = dex17
        ws['K40'].value = dex01
        ws['M40'].value = novan
        ws['N40'].value = colstop
        ws['O40'].value = colpux
        ws['P40'].value = retn
        ws['Q40'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 161:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E43'].value = total
        ws['F43'].value = dex08
        ws['G43'].value = dex07
        ws['H43'].value = dex03
        ws['I43'].value = dex05
        ws['J43'].value = dex17
        ws['K43'].value = dex01
        ws['M43'].value = novan
        ws['N43'].value = colstop
        ws['O43'].value = colpux
        ws['P43'].value = retn
        ws['Q43'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 162:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E44'].value = total
        ws['F44'].value = dex08
        ws['G44'].value = dex07
        ws['H44'].value = dex03
        ws['I44'].value = dex05
        ws['J44'].value = dex17
        ws['K44'].value = dex01
        ws['M44'].value = novan
        ws['N44'].value = colstop
        ws['O44'].value = colpux
        ws['P44'].value = retn
        ws['Q44'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 163:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E45'].value = total
        ws['F45'].value = dex08
        ws['G45'].value = dex07
        ws['H45'].value = dex03
        ws['I45'].value = dex05
        ws['J45'].value = dex17
        ws['K45'].value = dex01
        ws['M45'].value = novan
        ws['N45'].value = colstop
        ws['O45'].value = colpux
        ws['P45'].value = retn
        ws['Q45'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 164:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E46'].value = total
        ws['F46'].value = dex08
        ws['G46'].value = dex07
        ws['H46'].value = dex03
        ws['I46'].value = dex05
        ws['J46'].value = dex17
        ws['K46'].value = dex01
        ws['M46'].value = novan
        ws['N46'].value = colstop
        ws['O46'].value = colpux
        ws['P46'].value = retn
        ws['Q46'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 165:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E47'].value = total
        ws['F47'].value = dex08
        ws['G47'].value = dex07
        ws['H47'].value = dex03
        ws['I47'].value = dex05
        ws['J47'].value = dex17
        ws['K47'].value = dex01
        ws['M47'].value = novan
        ws['N47'].value = colstop
        ws['O47'].value = colpux
        ws['P47'].value = retn
        ws['Q47'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 166:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E48'].value = total
        ws['F48'].value = dex08
        ws['G48'].value = dex07
        ws['H48'].value = dex03
        ws['I48'].value = dex05
        ws['J48'].value = dex17
        ws['K48'].value = dex01
        ws['M48'].value = novan
        ws['N48'].value = colstop
        ws['O48'].value = colpux
        ws['P48'].value = retn
        ws['Q48'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 167:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E49'].value = total
        ws['F49'].value = dex08
        ws['G49'].value = dex07
        ws['H49'].value = dex03
        ws['I49'].value = dex05
        ws['J49'].value = dex17
        ws['K49'].value = dex01
        ws['M49'].value = novan
        ws['N49'].value = colstop
        ws['O49'].value = colpux
        ws['P49'].value = retn
        ws['Q49'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 168:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E50'].value = total
        ws['F50'].value = dex08
        ws['G50'].value = dex07
        ws['H50'].value = dex03
        ws['I50'].value = dex05
        ws['J50'].value = dex17
        ws['K50'].value = dex01
        ws['M50'].value = novan
        ws['N50'].value = colstop
        ws['O50'].value = colpux
        ws['P50'].value = retn
        ws['Q50'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 169:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E51'].value = total
        ws['F51'].value = dex08
        ws['G51'].value = dex07
        ws['H51'].value = dex03
        ws['I51'].value = dex05
        ws['J51'].value = dex17
        ws['K51'].value = dex01
        ws['M51'].value = novan
        ws['N51'].value = colstop
        ws['O51'].value = colpux
        ws['P51'].value = retn
        ws['Q51'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 170:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E52'].value = total
        ws['F52'].value = dex08
        ws['G52'].value = dex07
        ws['H52'].value = dex03
        ws['I52'].value = dex05
        ws['J52'].value = dex17
        ws['K52'].value = dex01
        ws['M52'].value = novan
        ws['N52'].value = colstop
        ws['O52'].value = colpux
        ws['P52'].value = retn
        ws['Q52'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 171:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E53'].value = total
        ws['F53'].value = dex08
        ws['G53'].value = dex07
        ws['H53'].value = dex03
        ws['I53'].value = dex05
        ws['J53'].value = dex17
        ws['K53'].value = dex01
        ws['M53'].value = novan
        ws['N53'].value = colstop
        ws['O53'].value = colpux
        ws['P53'].value = retn
        ws['Q53'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 172:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E54'].value = total
        ws['F54'].value = dex08
        ws['G54'].value = dex07
        ws['H54'].value = dex03
        ws['I54'].value = dex05
        ws['J54'].value = dex17
        ws['K54'].value = dex01
        ws['M54'].value = novan
        ws['N54'].value = colstop
        ws['O54'].value = colpux
        ws['P54'].value = retn
        ws['Q54'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 173:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E55'].value = total
        ws['F55'].value = dex08
        ws['G55'].value = dex07
        ws['H55'].value = dex03
        ws['I55'].value = dex05
        ws['J55'].value = dex17
        ws['K55'].value = dex01
        ws['M55'].value = novan
        ws['N55'].value = colstop
        ws['O55'].value = colpux
        ws['P55'].value = retn
        ws['Q55'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 201:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E58'].value = total
        ws['F58'].value = dex08
        ws['G58'].value = dex07
        ws['H58'].value = dex03
        ws['I58'].value = dex05
        ws['J58'].value = dex17
        ws['K58'].value = dex01
        ws['M58'].value = novan
        ws['N58'].value = colstop
        ws['O58'].value = colpux
        ws['P58'].value = retn
        ws['Q58'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 202:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E59'].value = total
        ws['F59'].value = dex08
        ws['G59'].value = dex07
        ws['H59'].value = dex03
        ws['I59'].value = dex05
        ws['J59'].value = dex17
        ws['K59'].value = dex01
        ws['M59'].value = novan
        ws['N59'].value = colstop
        ws['O59'].value = colpux
        ws['P59'].value = retn
        ws['Q59'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 203:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E60'].value = total
        ws['F60'].value = dex08
        ws['G60'].value = dex07
        ws['H60'].value = dex03
        ws['I60'].value = dex05
        ws['J60'].value = dex17
        ws['K60'].value = dex01
        ws['M60'].value = novan
        ws['N60'].value = colstop
        ws['O60'].value = colpux
        ws['P60'].value = retn
        ws['Q60'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 204:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E61'].value = total
        ws['F61'].value = dex08
        ws['G61'].value = dex07
        ws['H61'].value = dex03
        ws['I61'].value = dex05
        ws['J61'].value = dex17
        ws['K61'].value = dex01
        ws['M61'].value = novan
        ws['N61'].value = colstop
        ws['O61'].value = colpux
        ws['P61'].value = retn
        ws['Q61'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 205:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E62'].value = total
        ws['F62'].value = dex08
        ws['G62'].value = dex07
        ws['H62'].value = dex03
        ws['I62'].value = dex05
        ws['J62'].value = dex17
        ws['K62'].value = dex01
        ws['M62'].value = novan
        ws['N62'].value = colstop
        ws['O62'].value = colpux
        ws['P62'].value = retn
        ws['Q62'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 206:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E63'].value = total
        ws['F63'].value = dex08
        ws['G63'].value = dex07
        ws['H63'].value = dex03
        ws['I63'].value = dex05
        ws['J63'].value = dex17
        ws['K63'].value = dex01
        ws['M63'].value = novan
        ws['N63'].value = colstop
        ws['O63'].value = colpux
        ws['P63'].value = retn
        ws['Q63'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 207:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E64'].value = total
        ws['F64'].value = dex08
        ws['G64'].value = dex07
        ws['H64'].value = dex03
        ws['I64'].value = dex05
        ws['J64'].value = dex17
        ws['K64'].value = dex01
        ws['M64'].value = novan
        ws['N64'].value = colstop
        ws['O64'].value = colpux
        ws['P64'].value = retn
        ws['Q64'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 208:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E65'].value = total
        ws['F65'].value = dex08
        ws['G65'].value = dex07
        ws['H65'].value = dex03
        ws['I65'].value = dex05
        ws['J65'].value = dex17
        ws['K65'].value = dex01
        ws['M65'].value = novan
        ws['N65'].value = colstop
        ws['O65'].value = colpux
        ws['P65'].value = retn
        ws['Q65'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 301:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E67'].value = total
        ws['F67'].value = dex08
        ws['G67'].value = dex07
        ws['H67'].value = dex03
        ws['I67'].value = dex05
        ws['J67'].value = dex17
        ws['K67'].value = dex01
        ws['M67'].value = novan
        ws['N67'].value = colstop
        ws['O67'].value = colpux
        ws['P67'].value = retn
        ws['Q67'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 302:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E68'].value = total
        ws['F68'].value = dex08
        ws['G68'].value = dex07
        ws['H68'].value = dex03
        ws['I68'].value = dex05
        ws['J68'].value = dex17
        ws['K68'].value = dex01
        ws['M68'].value = novan
        ws['N68'].value = colstop
        ws['O68'].value = colpux
        ws['P68'].value = retn
        ws['Q68'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 303:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E69'].value = total
        ws['F69'].value = dex08
        ws['G69'].value = dex07
        ws['H69'].value = dex03
        ws['I69'].value = dex05
        ws['J69'].value = dex17
        ws['K69'].value = dex01
        ws['M69'].value = novan
        ws['N69'].value = colstop
        ws['O69'].value = colpux
        ws['P69'].value = retn
        ws['Q69'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 304:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E70'].value = total
        ws['F70'].value = dex08
        ws['G70'].value = dex07
        ws['H70'].value = dex03
        ws['I70'].value = dex05
        ws['J70'].value = dex17
        ws['K70'].value = dex01
        ws['M70'].value = novan
        ws['N70'].value = colstop
        ws['O70'].value = colpux
        ws['P70'].value = retn
        ws['Q70'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 305:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E71'].value = total
        ws['F71'].value = dex08
        ws['G71'].value = dex07
        ws['H71'].value = dex03
        ws['I71'].value = dex05
        ws['J71'].value = dex17
        ws['K71'].value = dex01
        ws['M71'].value = novan
        ws['N71'].value = colstop
        ws['O71'].value = colpux
        ws['P71'].value = retn
        ws['Q71'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 306:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E72'].value = total
        ws['F72'].value = dex08
        ws['G72'].value = dex07
        ws['H72'].value = dex03
        ws['I72'].value = dex05
        ws['J72'].value = dex17
        ws['K72'].value = dex01
        ws['M72'].value = novan
        ws['N72'].value = colstop
        ws['O72'].value = colpux
        ws['P72'].value = retn
        ws['Q72'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 307:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E73'].value = total
        ws['F73'].value = dex08
        ws['G73'].value = dex07
        ws['H73'].value = dex03
        ws['I73'].value = dex05
        ws['J73'].value = dex17
        ws['K73'].value = dex01
        ws['M73'].value = novan
        ws['N73'].value = colstop
        ws['O73'].value = colpux
        ws['P73'].value = retn
        ws['Q73'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 308:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E74'].value = total
        ws['F74'].value = dex08
        ws['G74'].value = dex07
        ws['H74'].value = dex03
        ws['I74'].value = dex05
        ws['J74'].value = dex17
        ws['K74'].value = dex01
        ws['M74'].value = novan
        ws['N74'].value = colstop
        ws['O74'].value = colpux
        ws['P74'].value = retn
        ws['Q74'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 309:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E75'].value = total
        ws['F75'].value = dex08
        ws['G75'].value = dex07
        ws['H75'].value = dex03
        ws['I75'].value = dex05
        ws['J75'].value = dex17
        ws['K75'].value = dex01
        ws['M75'].value = novan
        ws['N75'].value = colstop
        ws['O75'].value = colpux
        ws['P75'].value = retn
        ws['Q75'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 310:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E76'].value = total
        ws['F76'].value = dex08
        ws['G76'].value = dex07
        ws['H76'].value = dex03
        ws['I76'].value = dex05
        ws['J76'].value = dex17
        ws['K76'].value = dex01
        ws['M76'].value = novan
        ws['N76'].value = colstop
        ws['O76'].value = colpux
        ws['P76'].value = retn
        ws['Q76'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 311:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E77'].value = total
        ws['F77'].value = dex08
        ws['G77'].value = dex07
        ws['H77'].value = dex03
        ws['I77'].value = dex05
        ws['J77'].value = dex17
        ws['K77'].value = dex01
        ws['M77'].value = novan
        ws['N77'].value = colstop
        ws['O77'].value = colpux
        ws['P77'].value = retn
        ws['Q77'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 312:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E78'].value = total
        ws['F78'].value = dex08
        ws['G78'].value = dex07
        ws['H78'].value = dex03
        ws['I78'].value = dex05
        ws['J78'].value = dex17
        ws['K78'].value = dex01
        ws['M78'].value = novan
        ws['N78'].value = colstop
        ws['O78'].value = colpux
        ws['P78'].value = retn
        ws['Q78'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 313:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E79'].value = total
        ws['F79'].value = dex08
        ws['G79'].value = dex07
        ws['H79'].value = dex03
        ws['I79'].value = dex05
        ws['J79'].value = dex17
        ws['K79'].value = dex01
        ws['M79'].value = novan
        ws['N79'].value = colstop
        ws['O79'].value = colpux
        ws['P79'].value = retn
        ws['Q79'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 314:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E80'].value = total
        ws['F80'].value = dex08
        ws['G80'].value = dex07
        ws['H80'].value = dex03
        ws['I80'].value = dex05
        ws['J80'].value = dex17
        ws['K80'].value = dex01
        ws['M80'].value = novan
        ws['N80'].value = colstop
        ws['O80'].value = colpux
        ws['P80'].value = retn
        ws['Q80'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 315:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E81'].value = total
        ws['F81'].value = dex08
        ws['G81'].value = dex07
        ws['H81'].value = dex03
        ws['I81'].value = dex05
        ws['J81'].value = dex17
        ws['K81'].value = dex01
        ws['M81'].value = novan
        ws['N81'].value = colstop
        ws['O81'].value = colpux
        ws['P81'].value = retn
        ws['Q81'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 380:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E83'].value = total
        ws['F83'].value = dex08
        ws['G83'].value = dex07
        ws['H83'].value = dex03
        ws['I83'].value = dex05
        ws['J83'].value = dex17
        ws['K83'].value = dex01
        ws['M83'].value = novan
        ws['N83'].value = colstop
        ws['O83'].value = colpux
        ws['P83'].value = retn
        ws['Q83'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 441:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E90'].value = total
        ws['F90'].value = dex08
        ws['G90'].value = dex07
        ws['H90'].value = dex03
        ws['I90'].value = dex05
        ws['J90'].value = dex17
        ws['K90'].value = dex01
        ws['M90'].value = novan
        ws['N90'].value = colstop
        ws['O90'].value = colpux
        ws['P90'].value = retn
        ws['Q90'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 443:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E91'].value = total
        ws['F91'].value = dex08
        ws['G91'].value = dex07
        ws['H91'].value = dex03
        ws['I91'].value = dex05
        ws['J91'].value = dex17
        ws['K91'].value = dex01
        ws['M91'].value = novan
        ws['N91'].value = colstop
        ws['O91'].value = colpux
        ws['P91'].value = retn
        ws['Q91'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 444:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E92'].value = total
        ws['F92'].value = dex08
        ws['G92'].value = dex07
        ws['H92'].value = dex03
        ws['I92'].value = dex05
        ws['J92'].value = dex17
        ws['K92'].value = dex01
        ws['M92'].value = novan
        ws['N92'].value = colstop
        ws['O92'].value = colpux
        ws['P92'].value = retn
        ws['Q92'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 445:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E93'].value = total
        ws['F93'].value = dex08
        ws['G93'].value = dex07
        ws['H93'].value = dex03
        ws['I93'].value = dex05
        ws['J93'].value = dex17
        ws['K93'].value = dex01
        ws['M93'].value = novan
        ws['N93'].value = colstop
        ws['O93'].value = colpux
        ws['P93'].value = retn
        ws['Q93'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 446:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E94'].value = total
        ws['F94'].value = dex08
        ws['G94'].value = dex07
        ws['H94'].value = dex03
        ws['I94'].value = dex05
        ws['J94'].value = dex17
        ws['K94'].value = dex01
        ws['M94'].value = novan
        ws['N94'].value = colstop
        ws['O94'].value = colpux
        ws['P94'].value = retn
        ws['Q94'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 447:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E95'].value = total
        ws['F95'].value = dex08
        ws['G95'].value = dex07
        ws['H95'].value = dex03
        ws['I95'].value = dex05
        ws['J95'].value = dex17
        ws['K95'].value = dex01
        ws['M95'].value = novan
        ws['N95'].value = colstop
        ws['O95'].value = colpux
        ws['P95'].value = retn
        ws['Q95'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 449:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E96'].value = total
        ws['F96'].value = dex08
        ws['G96'].value = dex07
        ws['H96'].value = dex03
        ws['I96'].value = dex05
        ws['J96'].value = dex17
        ws['K96'].value = dex01
        ws['M96'].value = novan
        ws['N96'].value = colstop
        ws['O96'].value = colpux
        ws['P96'].value = retn
        ws['Q96'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 450:
        total = int(input('What is the total number of package? '))
        dex08 = int(input('How many DEX08 has he got? '))
        dex07 = int(input('How many DEX07 has he got? '))
        dex03 = int(input('How many DEX03 has he got? '))
        dex05 = int(input('How many DEX05 has he got? '))
        dex17 = int(input('How many DEX17 has he got? '))
        dex01 = int(input('How many DEX01 has he got? '))
        novan = int(input('How many NO VANS? '))
        colstop = int(input('How many collection stops? '))
        colpux = int(input('How many Failed stop PUX? '))
        retn = input('What time driver return to work? ')
        cmt = input('Is there any comment? ')


        wb = load_workbook('Returns.xlsx')
        ws = wb.active
        ws['E97'].value = total
        ws['F97'].value = dex08
        ws['G97'].value = dex07
        ws['H97'].value = dex03
        ws['I97'].value = dex05
        ws['J97'].value = dex17
        ws['K97'].value = dex01
        ws['M97'].value = novan
        ws['N97'].value = colstop
        ws['O97'].value = colpux
        ws['P97'].value = retn
        ws['Q97'].value = cmt

        wb.save('Returns.xlsx')

    elif route == 00:
        print('The Return Analysis work sheet has been updated. Please update the driver name, mobile number, postcodes and the route number manually in the sheet')

        break
    else:
        print('The route number is not reconised. Please contact Mohamed for assistance.')
